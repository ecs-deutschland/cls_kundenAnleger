import pyautogui
import requests
from bs4 import BeautifulSoup as bs
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime
from openpyxl import load_workbook

def excelToPython(dateipfad = '_kundendaten_SID.xlsx', tabellenname = 'tab',ZeilenBeginn = 2, ZeilenEnde = 90):   # 2, da die erste Zeile die Beschreibungen sind
    DATEIPFAD = dateipfad
    TABELLENNAME = tabellenname
    wb = load_workbook(filename = DATEIPFAD)
    tabelle = wb[TABELLENNAME]

    Ergebnisliste = []  # [ [firmenname,adresse,ort,bundesland,plz,vorname,nachname,email], [firmenname...]]
                        # wobei 1. Eintrag = 1. Kunde, 2. Eintrag = 2.Kunde ...

    for i in range(ZeilenBeginn,ZeilenEnde):
        kundenEintraege = []
        if tabelle["B" + str(i)].value != None:
            kundenEintraege.append(tabelle['B' + str(i)].value) # firmenname
            kundenEintraege.append(tabelle['D' + str(i)].value) # adresse
            kundenEintraege.append(tabelle['E' + str(i)].value) # plz
            kundenEintraege.append(tabelle['F' + str(i)].value) # ort
            kundenEintraege.append(tabelle['G' + str(i)].value) # vorname
            kundenEintraege.append(tabelle['H' + str(i)].value) # nachnacme
            kundenEintraege.append(tabelle['J' + str(i)].value) # email
        Ergebnisliste.append(kundenEintraege)
    return Ergebnisliste

def pythonToExcel(dateipfad = '_kundendaten_SID.xlsx', tabellenname = 'tab', zeile=2, spalte='Z', inhalt="testInhalt"):
    workbook = load_workbook(dateipfad, read_only = False)
    tabelle = workbook[tabellenname]
    zellenwert = spalte + str(zeile)
    zelle = tabelle[zellenwert]
    zelle.value = inhalt
    workbook.save(dateipfad)
    print("In Zelle: '" + spalte + str(zeile) + "' Inhalt: '" + inhalt + "' eingetragen (Dokument: '" + dateipfad + "').")

def zeichenAendern(eingabe):
    zeichenVerbessertesWort = eingabe
    if "ä" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ä","ae")
    if "ü" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ü","ue")
    if "ö" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ö","oe")
    if "Ä" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ä","Ae")
    if "Ü" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ü","Ue")
    if "Ö" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ö","Oe")
    if "ß" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ß","ss")
    return zeichenVerbessertesWort

def umlauteAendern(string):
    finalerString = ""
    for i in string:
        finalerString += (zeichenAendern(i))
    return finalerString

class Scraper():

    def __init__(self):
        self.now = datetime.now()
        self.startzeit = self.now.strftime("%H:%M:%S")
        url = "https://www.central-license-server.com/poeticWeb/session/login.htm"
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(url)
        self.wait = WebDriverWait(self.driver, 3)
        print("CLS Eintragung um " + self.startzeit + " begonnen!")

    def login(self, user = "HolfeldM", password = "Schweden21!"):
        pyautogui.write(user)
        pyautogui.press("tab")
        pyautogui.write(password)
        time.sleep(1)
        self.clickByLinkText("Anmelden")
        print("Eingeloggt als " + user + "...")
        pyautogui.sleep(1)

    def firmenSucheOeffnen(self):
        self.clickByLinkText("Aktivierung")
        self.clickByLinkText("Lizenzschlüssel generieren und herunterladen")
        self.clickByXPath('//a[img/@src="/poeticWeb/images/find_button_small.gif"]')

    def kundeAnlegen(self, firmenname = " ", adresse = " ", postleitzahl = " ", ort = " ", vorname = " ", nachname = " ", email = " "):
        kundeGefunden = self.namensTest(firmenname)
        print("Endergebnis 'self.kundeGefunden()' = " + str(kundeGefunden))
        if kundeGefunden == False: # Kunde nicht vorhanden
            self.clickByLinkText("[Firma hinzufügen]")
            now = datetime.now()
            print("Firmeneintragung: '" + firmenname + "' um " + now.strftime("%H:%M:%S") + " begonnen.")
            time.sleep(1)
            pubID = self.getPubID()
            print("Neue PUB-ID von '" + firmenname + ": " + str(pubID) + ".")
            pyautogui.write(firmenname, 0.1)
            print("Name eingetragen: '" + str(firmenname) + "'")
            time.sleep(0.5)
            self.clickByXPath("//select/option[@value='DEU']")
            time.sleep(1)
            self.multiPress(3)
            pyautogui.write(adresse, 0.1)
            print("Adresse eingetragen: '" + str(adresse) + "'")
            pyautogui.press('tab',presses=3, interval=0.2)
            pyautogui.write(str(postleitzahl), 0.1)
            print("Postleitzahl eingetragen: '" + str(postleitzahl) + "'")
            pyautogui.press("tab")
            pyautogui.write(ort, 0.1)
            print("Ort eingetragen: '" + str(ort) + "'")
            pyautogui.press('tab',presses=3, interval=0.2)
            pyautogui.write(vorname, 0.1)
            print("Vorname eingetragen: '" + str(vorname) + "'")
            pyautogui.press('tab',presses=2,interval=0.2)
            pyautogui.write(nachname, 0.1)
            print("Nachname eingetragen: '" + str(nachname) + "'")
            self.multiPress(6)
            self.enterEmail(email)
            print("Email eingetragen: '" + str(email) + "'")
            self.multiPress(5)
            self.countdown(5)              # zählt Countdown runter bevor gespeichert wird
            pyautogui.press("enter")        # speichern drücken
            return pubID
        else:
            print("Stringpattern: '" + firmenname +  "' im System gefunden.")
            self.clearField()

    def pubToExcel(self):
        pass

    def getPubID(self):
        pubID = self.driver.find_element_by_xpath('//input[@name="publisherXRefNumber"]').get_attribute("value")
        return pubID

    def enterEmail(self, mail):
        atPosition = self.findATinMail(mail)
        pyautogui.write(mail[0:atPosition], 0.1)
        pyautogui.hotkey('ctrl','alt','q')
        pyautogui.write(mail[atPosition:len(mail)], 0.1)

    def findATinMail(self, mail):
        return mail.rfind("@")

    def namensTest(self, name):
        gefunden = False
        while len(name.split()) > 7:               # name.split() zählt Wörter des Strings
            print("Längenmessung: " + str(len(name.split())))
            kundeGefunden = self.kundeSuchen(name)
            if kundeGefunden == True:  # wurde Kunde  gefunden
                self.clearField()
                self.clickByLinkText("Firmen suchen")
                gefunden = True
                break
            else:                               # Kunde wurde nicht gefunden
                self.clearField()
                print("Eingabefeld geleert!")
                #kundeTest(name[:name.rfind(' ')])    # schaut letzte Leerzeichenstelle und splittet String dort, prüft erneut
            name = name[:name.rfind(' ')]
        return gefunden

    def kundeSuchen(self, name = "--- testName ---", pubID = "00000"):
        nameEnter = name + "%"
        time.sleep(1)
        pyautogui.write(nameEnter, 0.1)
        print("Firma: '" + nameEnter + "' im Suchfeld eingegeben.")
        self.clickByLinkText("Firmen suchen")
        return self.checkExistence("Auswählen", name)

    def checkExistence(self, text = "Auswählen", name = ""):
        try:
            elements = self.wait.until(EC.presence_of_all_elements_located((By.PARTIAL_LINK_TEXT, text)))
            print("Unternehmen '" + name + "' bereits eingetragen")
            return True
        except selenium.common.exceptions.TimeoutException:
            print("Unternehmen '" +  name + "' noch nicht eingetragen")
            return False

    def clearField(self):
        pyautogui.hotkey('ctrl','a')
        time.sleep(0.5)
        pyautogui.press('delete')

    def iframe(self, frame = " "):
        self.driver.switch_to.frame(frame)

    def multiPress(self, anzahl = 2, taste = 'tab'):
        while 0 < anzahl:
            pyautogui.press(taste, interval = 0.1)
            anzahl -= 1

    def clickByClass(self, klasse):
        wait = WebDriverWait(self.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, klasse)))
        element.click()

    def clickByLinkText(self, linktext):
        wait = WebDriverWait(self.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, linktext)))
        element.click()

    def clickByID(self, id):
        wait = WebDriverWait(self.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.ID, id)))
        element.click()

    def clickByName(self, name):
        wait = WebDriverWait(self.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.NAME, name)))
        element.click()

    def clickByXPath(self, XPath):
        self.driver.find_element_by_xpath(XPath).click()

    def findByXPath(self, XPath):
        return self.driver.find_elements_by_xpath(XPath).text

    def countdown(self, zeit):
        print("EINTRAGUNG WIRD IN" +  str(zeit) + " SEKUNDEN AUSGEFÜHRT!")
        while(0 < zeit):
            print(zeit)
            time.sleep(1)
            zeit -= 1

data = excelToPython()

AnlegeObjekt = Scraper()
AnlegeObjekt.login()
AnlegeObjekt.firmenSucheOeffnen()

for i in range(len(data)):
    ''' Diese Schleife geht alle Excel-Einträge durch, überprüft ob eine Firma vorhanden ist und trägt sie ein, falls nötig '''
    print("_______________ EINTRAG " + str(i+1) + " _______________")
    print("Trage Kunde: '" + str(data[i][0]) + "' mit: '" + str(data[i][1]) + "', '" + str(data[i][2]) + "', '" + str(data[i][3])  + "', '" + str(data[i][4]) + "', '" + str(data[i][5]) + "', '" + str(data[i][6]) + "' ein!")
    pubID = AnlegeObjekt.kundeAnlegen(umlauteAendern(data[i][0]),umlauteAendern(data[i][1]),data[i][2],umlauteAendern(data[i][3]),umlauteAendern(data[i][4]),umlauteAendern(data[i][5]),data[i][6])
    if len(str(pubID)) <= 4:
        print(type(pubID))
    else:
        pythonToExcel(zeile = i+2, spalte = "M", inhalt=pubID)
        time.sleep(1)
        AnlegeObjekt.clickByXPath('//a[img/@src="/poeticWeb/images/find_button_small.gif"]')
